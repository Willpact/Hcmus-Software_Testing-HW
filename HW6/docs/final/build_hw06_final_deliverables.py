"""Build HW06 final workbook and PDF from preserved local artifacts.

The script only reads final testcase/run/defect/report artifacts and never reads
or writes a runtime environment, SUT, test data, or Git state.
"""

from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    KeepTogether,
    Image as PdfImage,
    ListFlowable,
    ListItem,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table as PdfTable,
    TableStyle,
)


ROOT = Path(__file__).resolve().parents[2]
FINAL = ROOT / "docs" / "final"
XLSX_PATH = FINAL / "HW06-Test-Cases-and-Summary.xlsx"
PDF_PATH = FINAL / "HW06-MAIN-REPORT.pdf"
REPORT_PATH = FINAL / "HW06-MAIN-REPORT.md"

NAVY = "17365D"
BLUE = "D9EAF7"
GREEN = "E2F0D9"
AMBER = "FFF2CC"
RED = "FCE4D6"
GRAY = "F2F2F2"


def load_json(relative: str):
    return json.loads((ROOT / relative).read_text(encoding="utf-8-sig"))


def flatten(value) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        return "; ".join(str(item) for item in value)
    return str(value)


def compact(value: str, limit: int = 500) -> str:
    value = re.sub(r"\s+", " ", value or "").strip()
    return value if len(value) <= limit else value[: limit - 1] + "…"


def load_cases_and_outcomes():
    cases = []
    for filename in (
        "api-01-reset-password.json",
        "api-02-checkout.json",
        "api-03-import-products.json",
    ):
        payload = load_json(f"test-cases/final/{filename}")
        cases.extend(payload["cases"])

    run1 = load_json("test-results/hw06/run-001/case-accounting.json")
    run2 = load_json("test-results/hw06/run-002/case-accounting.json")
    outcomes = {}
    for row in run1["cases"]:
        outcomes[row["case_id"]] = ("run-001", row)
    # run-002 is the newest evidence for its selected identities.
    for row in run2["cases"]:
        outcomes[row["case_id"]] = ("run-002", row)
    return cases, run1, run2, outcomes


def build_workbook():
    cases, run1, run2, outcomes = load_cases_and_outcomes()
    summary = load_json("test-cases/final/cross-api-final-summary.json")
    audit_summary = load_json("test-cases/audited/cross-api-summary.json")["total_summary"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    headers = [
        "Case ID", "API", "Source", "Human Review", "Final Disposition",
        "Requirement IDs", "Primary Technique", "Endpoint", "Title", "Objective",
        "Preconditions", "Request Variation", "Expected Business Result", "Expected State",
        "Execution Mode", "External Verification", "Latest Evidence Run", "Business/State Result",
        "Classification", "HTTP Status", "Root Cluster(s)",
    ]
    ws.merge_cells("A1:U1")
    ws["A1"] = "HW06 — Final Test Cases (93 executable cases)"
    ws["A1"].font = Font(name="Aptos Display", bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A2:U2")
    ws["A2"] = "Nguồn: test-cases/final/*.json; kết quả execution lấy từ preserved run-001/run-002. Không chứa credential hoặc test-data secret."
    ws["A2"].font = Font(italic=True, color="404040")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 30
    ws.append([])
    ws.append(headers)

    for entry in cases:
        tc = entry["test_case"]
        run_name, observed = outcomes.get(entry["case_id"], ("", {}))
        classification = observed.get("classification", observed.get("preliminary_classification", ""))
        clusters = flatten(observed.get("root_clusters", []))
        ws.append([
            entry["case_id"], entry["api_id"], entry["source"], entry["human_review_status"],
            entry["final_disposition"], flatten(tc.get("requirement_ids")), tc.get("primary_technique", ""),
            tc.get("endpoint", ""), tc.get("title", ""), compact(tc.get("objective", "")),
            compact(flatten(tc.get("preconditions"))), compact(tc.get("request", {}).get("body_variation", "")),
            compact(tc.get("expected_business_result", "")), compact(tc.get("expected_state", "")),
            entry.get("execution_mode", ""), entry.get("external_verification", ""), run_name,
            observed.get("result", ""), classification, observed.get("observed_status_code", ""), clusters,
        ])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[4]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[4].height = 34
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    widths = [16, 11, 16, 14, 20, 26, 19, 28, 30, 42, 40, 34, 40, 40, 30, 30, 18, 22, 28, 13, 25]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.auto_filter.ref = f"A4:U{ws.max_row}"
    tab = Table(displayName="FinalTestCases", ref=f"A4:U{ws.max_row}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    ws.add_table(tab)
    ws.conditional_formatting.add(f"R5:R{ws.max_row}", CellIsRule(operator="equal", formula=['"FAIL"'], fill=PatternFill("solid", fgColor=RED)))
    ws.conditional_formatting.add(f"R5:R{ws.max_row}", CellIsRule(operator="equal", formula=['"PASS"'], fill=PatternFill("solid", fgColor=GREEN)))

    summary_ws = wb.create_sheet("Test Summary")
    summary_ws.sheet_view.showGridLines = False
    summary_ws.merge_cells("A1:F1")
    summary_ws["A1"] = "HW06 — Test Summary"
    summary_ws["A1"].font = Font(name="Aptos Display", bold=True, size=16, color="FFFFFF")
    summary_ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    summary_ws["A1"].alignment = Alignment(horizontal="center")
    summary_ws.row_dimensions[1].height = 28
    summary_ws["A3"] = "Test design inventory"
    summary_ws["A3"].font = Font(bold=True, color="FFFFFF")
    summary_ws["A3"].fill = PatternFill("solid", fgColor="1F4E78")
    summary_ws.merge_cells("A3:F3")
    design_rows = [
        ["Metric", "Value", "Source / interpretation"],
        ["Selected APIs", 3, "API-01 Password Reset; API-02 Checkout; API-03 Import Products"],
        ["AI-generated candidates", summary["totals"]["raw_ai_generated"], "cross-api-final-summary.json"],
        ["AI Audit VALID", audit_summary["valid"], "test-cases/audited/cross-api-summary.json"],
        ["AI Audit INVALID", audit_summary["invalid"], "test-cases/audited/cross-api-summary.json"],
        ["AI Audit INCOMPLETE", audit_summary["incomplete"], "test-cases/audited/cross-api-summary.json"],
        ["AI corrected executable", summary["totals"]["ai_corrected_executable"], "cross-api-final-summary.json"],
        ["Student-added approved", summary["totals"]["student_added_approved"], "cross-api-final-summary.json"],
        ["Final executable test cases", "=COUNTA('Test Cases'!A5:A97)", "Formula based on Test Cases sheet"],
        ["Deferred requirement-gap candidates", summary["totals"]["deferred"], "Excluded from final executable suite"],
        ["Invalid removed candidates", summary["totals"]["invalid_removed"], "Excluded from final executable suite"],
        ["Confirmed product defects", 9, "docs/defects/evidence-matrix.md"],
        ["Product-defect evidence testcase records", 38, "Evidence records for 9 defects; not 38 defects"],
    ]
    for row in design_rows:
        summary_ws.append(row)
    for cell in summary_ws[4]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    summary_ws["A20"] = "Genuine execution summary — do not combine runs"
    summary_ws["A20"].font = Font(bold=True, color="FFFFFF")
    summary_ws["A20"].fill = PatternFill("solid", fgColor="1F4E78")
    summary_ws.merge_cells("A20:F20")
    execution_rows = [
        ["Run", "Testcase scope", "Passed", "Failed", "Blocked", "External pending", "Notes"],
        ["run-001", 93, 27, 38, 27, 1, "Full preserved execution; 0 Newman assertion failures"],
        ["run-002", 37, 15, 21, 1, 0, "Targeted corrective rerun; 0 Newman assertion failures"],
    ]
    for row in execution_rows:
        summary_ws.append(row)
    for cell in summary_ws[21]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    summary_ws["A26"] = "Important interpretation"
    summary_ws["A26"].font = Font(bold=True, color="FFFFFF")
    summary_ws["A26"].fill = PatternFill("solid", fgColor="1F4E78")
    summary_ws.merge_cells("A26:F26")
    summary_ws.merge_cells("A27:F28")
    summary_ws["A27"] = "Business/state FAIL is established by approved oracle and external verification. It is not a Newman assertion failure. The two runs have different scopes, so their PASS/FAIL/BLOCKED counts must not be summed."
    summary_ws["A27"].alignment = Alignment(wrap_text=True, vertical="top")
    summary_ws["A27"].fill = PatternFill("solid", fgColor="FFF2CC")
    for column, width in {"A": 34, "B": 18, "C": 18, "D": 18, "E": 18, "F": 18, "G": 54}.items():
        summary_ws.column_dimensions[column].width = width
    for row in summary_ws.iter_rows(min_row=4, max_row=summary_ws.max_row, min_col=1, max_col=7):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    summary_ws.freeze_panes = "A4"

    defects_ws = wb.create_sheet("Defects")
    defects_ws.sheet_view.showGridLines = False
    defects_ws.merge_cells("A1:F1")
    defects_ws["A1"] = "HW06 — Confirmed Product Defects"
    defects_ws["A1"].font = Font(name="Aptos Display", bold=True, size=16, color="FFFFFF")
    defects_ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    defects_ws["A1"].alignment = Alignment(horizontal="center")
    defects_ws.append([])
    defect_headers = ["Defect ID", "API", "Title", "Primary Case", "Evidence", "GitHub Issue"]
    defects_ws.append(defect_headers)
    titles = [
        ("DEF-01", "Checkout API", "Checkout tin tưởng total_amount do client gửi lên", "API02-AI-002", "Request/response + state", "Issue #393"),
        ("DEF-02", "Checkout API", "Giỏ hàng không được xóa sau khi checkout thành công", "API02-AI-014", "Request/response + state", "Issue #394"),
        ("DEF-03", "Checkout API", "Checkout không bắt buộc đúng Bearer authorization scheme", "API02-AI-022", "Request/response", "Issue #395"),
        ("DEF-04", "Import Products API", "API import sản phẩm chấp nhận giá không dương", "API03-AI-009", "Request/response + state", "Issue #396"),
        ("DEF-05", "Import Products API", "Import sản phẩm không đảm bảo tính nguyên tử", "API03-AI-017", "Request/response + state", "Issue #397"),
        ("DEF-06", "Import Products API", "API import sản phẩm không kiểm tra quyền Admin", "API03-AI-026", "Request/response + state", "Issue #398"),
        ("DEF-07", "Password Reset API", "Reset mật khẩu vẫn thành công khi thiếu mật khẩu mới", "API01-AI-007", "Request/response + state", "Issue #399"),
        ("DEF-08", "Password Reset API", "Quy tắc độ mạnh mật khẩu không được kiểm tra khi reset", "API01-AI-018", "Request/response + state", "Issue #400"),
        ("DEF-09", "Password Reset API", "Mật khẩu mới được lưu dưới dạng plaintext", "API01-AI-035", "Request/response + state", "Issue #401"),
    ]
    for row in titles:
        defects_ws.append(row)
    for cell in defects_ws[3]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for row in defects_ws.iter_rows(min_row=4, max_row=defects_ws.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column, width in {"A": 14, "B": 26, "C": 64, "D": 20, "E": 30, "F": 16}.items():
        defects_ws.column_dimensions[column].width = width
    defects_ws.freeze_panes = "A4"
    defect_table = Table(displayName="ConfirmedDefects", ref=f"A3:F{defects_ws.max_row}")
    defect_table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    defects_ws.add_table(defect_table)

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.row not in (1, 2, 3, 4, 20, 21, 26) and cell.font == Font():
                    cell.font = Font(name="Aptos", size=10)
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.page_margins.left = 0.3
        sheet.page_margins.right = 0.3
        sheet.page_margins.top = 0.5
        sheet.page_margins.bottom = 0.5

    wb.save(XLSX_PATH)
    return len(cases)


def _inline(text: str) -> str:
    text = text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    text = re.sub(r"\[([^\]]+)\]\((https?://[^)]+)\)", r"<link href='\2' color='#1F4E78'>\1</link>", text)
    text = re.sub(r"`([^`]+)`", r"<font name='Courier'>\1</font>", text)
    text = re.sub(r"\*\*(.+?)\*\*", r"<b>\1</b>", text)
    return text


def _table_from_markdown(lines, styles):
    rows = []
    for line in lines:
        if re.match(r"^\|\s*[-:]+", line):
            continue
        cells = [cell.strip() for cell in line.strip().strip("|").split("|")]
        rows.append([Paragraph(_inline(cell), styles["TableCell"]) for cell in cells])
    width = 17.0 * cm
    col_count = max(len(row) for row in rows)
    table = PdfTable(rows, colWidths=[width / col_count] * col_count, repeatRows=1, hAlign="LEFT")
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Arial-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D9E2F3")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7FBFF")]),
    ]))
    return table


def build_pdf():
    pdfmetrics.registerFont(TTFont("Arial", r"C:\Windows\Fonts\arial.ttf"))
    pdfmetrics.registerFont(TTFont("Arial-Bold", r"C:\Windows\Fonts\arialbd.ttf"))
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="TitleVN", parent=styles["Title"], fontName="Arial-Bold", fontSize=19, leading=24, textColor=colors.HexColor("#17365D"), alignment=TA_CENTER, spaceAfter=14))
    styles.add(ParagraphStyle(name="H1VN", parent=styles["Heading1"], fontName="Arial-Bold", fontSize=14, leading=18, textColor=colors.HexColor("#17365D"), spaceBefore=12, spaceAfter=7))
    styles.add(ParagraphStyle(name="H2VN", parent=styles["Heading2"], fontName="Arial-Bold", fontSize=11, leading=14, textColor=colors.HexColor("#1F4E78"), spaceBefore=9, spaceAfter=5))
    styles.add(ParagraphStyle(name="BodyVN", parent=styles["BodyText"], fontName="Arial", fontSize=9, leading=12.5, spaceAfter=5))
    styles.add(ParagraphStyle(name="TableCell", parent=styles["BodyText"], fontName="Arial", fontSize=7.5, leading=9.5))
    styles.add(ParagraphStyle(name="FooterVN", parent=styles["BodyText"], fontName="Arial", fontSize=8, textColor=colors.HexColor("#666666")))
    story = [Paragraph("Báo cáo chính HW06 — API Testing", styles["TitleVN"]), Paragraph("Nguồn tổng hợp: artifact thật trong workspace HW6; không suy diễn kết quả CI chưa có.", styles["FooterVN"]), Spacer(1, 0.3 * cm)]
    lines = REPORT_PATH.read_text(encoding="utf-8").splitlines()
    i = 1 if lines and lines[0].startswith("# ") else 0
    paragraph_buffer = []
    bullets = []

    def flush_paragraph():
        nonlocal paragraph_buffer
        if paragraph_buffer:
            story.append(Paragraph(_inline(" ".join(paragraph_buffer)), styles["BodyVN"]))
            paragraph_buffer = []

    def flush_bullets():
        nonlocal bullets
        if bullets:
            story.append(ListFlowable([ListItem(Paragraph(_inline(item), styles["BodyVN"])) for item in bullets], bulletType="bullet", leftIndent=14))
            bullets = []

    while i < len(lines):
        line = lines[i]
        image_match = re.match(r"^!\[([^\]]+)\]\(([^)]+)\)$", line)
        if image_match:
            flush_paragraph(); flush_bullets()
            image_path = (REPORT_PATH.parent / image_match.group(2)).resolve()
            figure = PdfImage(str(image_path))
            scale = min((5.5 * cm) / figure.imageWidth, (18.0 * cm) / figure.imageHeight)
            figure.drawWidth = figure.imageWidth * scale
            figure.drawHeight = figure.imageHeight * scale
            figure.hAlign = "CENTER"
            story.append(figure)
            story.append(Spacer(1, 0.18 * cm))
            i += 1
            continue
        elif line.startswith("|"):
            flush_paragraph(); flush_bullets()
            table_lines = []
            while i < len(lines) and lines[i].startswith("|"):
                table_lines.append(lines[i]); i += 1
            story.append(_table_from_markdown(table_lines, styles)); story.append(Spacer(1, 0.18 * cm)); continue
        if line.startswith("## "):
            flush_paragraph(); flush_bullets()
            if line[3:].startswith("Figure — "):
                story.append(PageBreak())
            story.append(Paragraph(_inline(line[3:]), styles["H1VN"]))
        elif line.startswith("### "):
            flush_paragraph(); flush_bullets(); story.append(Paragraph(_inline(line[4:]), styles["H2VN"]))
        elif re.match(r"^- ", line):
            flush_paragraph(); bullets.append(line[2:])
        elif not line.strip():
            flush_paragraph(); flush_bullets()
        else:
            paragraph_buffer.append(line.strip())
        i += 1
    flush_paragraph(); flush_bullets()

    def footer(canvas, doc):
        canvas.saveState()
        canvas.setFont("Arial", 8)
        canvas.setFillColor(colors.HexColor("#666666"))
        canvas.drawString(1.8 * cm, 1.1 * cm, "HW06 API Testing — final report")
        canvas.drawRightString(19.2 * cm, 1.1 * cm, f"Trang {doc.page}")
        canvas.restoreState()

    doc = SimpleDocTemplate(str(PDF_PATH), pagesize=A4, leftMargin=1.8 * cm, rightMargin=1.8 * cm, topMargin=1.3 * cm, bottomMargin=1.5 * cm, title="HW06 Main Report")
    doc.build(story, onFirstPage=footer, onLaterPages=footer)


if __name__ == "__main__":
    count = build_workbook()
    build_pdf()
    print(f"TEST_CASES={count}")
    print(f"XLSX={XLSX_PATH}")
    print(f"PDF={PDF_PATH}")
